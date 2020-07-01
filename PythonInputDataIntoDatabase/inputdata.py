from openpyxl.reader.excel import load_workbook
import mysql.connector
import datetime

cnx = mysql.connector.connect(user='root', password='34101130',
								host='localhost', database='crmdb')
cursor = cnx.cursor()

workbook = load_workbook("CRMdatabase.xlsx", data_only=True)
sheetnames = workbook.get_sheet_names()

gender = {"M": "Male", "F": "Female"}

print (sheetnames)
worksheet = workbook.get_sheet_by_name(sheetnames[0]) # store
add_store = ("INSERT INTO store "
			 "(storeID, storeName, storeAddress, storePhoneNumber, storeCity, storeState) "
			 "VALUES (%s, %s, %s, %s, %s, %s)")
			 
for row_index in range(2, worksheet.max_row + 1):
	data_store = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value, worksheet.cell(row_index, 3).value, worksheet.cell(row_index, 4).value,
				 worksheet.cell(row_index, 5).value, worksheet.cell(row_index, 6).value,)
	cursor.execute(add_store, data_store)


worksheet = workbook.get_sheet_by_name(sheetnames[4]) # vehicle
add_vehicle = ("INSERT INTO vehicle "
			 "(vehicleID, makename, model, series, seriesYear, priceNew, engineSize, fuelSystem, tankCapacity, power, seatingCapacity, standardTransmission, bodyType, wheelDrive, wheelBase) "
			 "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")

for row_index in range(2, worksheet.max_row + 1):
	data_vehicle = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value, worksheet.cell(row_index, 3).value, worksheet.cell(row_index, 4).value,
				 worksheet.cell(row_index, 5).value, worksheet.cell(row_index, 6).value, float(worksheet.cell(row_index, 7).value.strip('L')), worksheet.cell(row_index, 8).value,
				 float(worksheet.cell(row_index, 9).value.strip('L')), float(worksheet.cell(row_index, 10).value.strip('Kw')), worksheet.cell(row_index, 11).value, worksheet.cell(row_index, 12).value,
				 worksheet.cell(row_index, 13).value, worksheet.cell(row_index, 14).value, float(worksheet.cell(row_index, 15).value.strip('mm')))
	cursor.execute(add_vehicle, data_vehicle)

def get_formatted_value(value):
	if type(value) is str:
		return gender[value.strip(' ')]
	else:
		return 'Other'

worksheet = workbook.get_sheet_by_name(sheetnames[1]) # Customer
add_customer = ("INSERT INTO customer "
			 "(customerID, customerFirstName, customerLastName, customerDOB, customerAddress, customerOccupation, customerGender) "
			 "VALUES (%s, %s, %s, %s, %s, %s, %s)")
print (worksheet.max_row)
for row_index in range(2, 109):
	data_customer = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value, worksheet.cell(row_index, 3).value, worksheet.cell(row_index, 4).value,
				 worksheet.cell(row_index, 5).value, worksheet.cell(row_index, 6).value, get_formatted_value(worksheet.cell(row_index, 7).value))
	
	print (data_customer)
	cursor.execute(add_customer, data_customer)

worksheet = workbook.get_sheet_by_name(sheetnames[2]) # customer phonenumber
add_customerphonenumber = ("INSERT INTO customerphonenumber "
						   "(customerID, phoneNumber) "
						   "VALUES (%s, %s)")
for row_index in range(2, 109):
	data_customerphonenumber = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value)
	cursor.execute(add_customerphonenumber, data_customerphonenumber)
						   
worksheet = workbook.get_sheet_by_name(sheetnames[3]) # employeerole
add_employeerole = ("INSERT INTO employeerole "
					"(roleID, roleName, roleAccessLevel) "
					"VALUES (%s, %s, %s)")
for row_index in range(2, 4):
	data_employeerole = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value, worksheet.cell(row_index, 3).value)
	cursor.execute(add_employeerole, data_employeerole)
					
worksheet = workbook.get_sheet_by_name(sheetnames[5]) # employee
add_employee = ("INSERT INTO employee "
					"(employeeID, employeeFirstName, employeeLastName, employeeDOB, employeeAddress, employeeCity, employeeState, employeeEmail, employeePhoneNumber, roleID, employeePassword) "
					"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
for row_index in range(2, 5):
	data_employee = (worksheet.cell(row_index, 1).value, worksheet.cell(row_index, 2).value, worksheet.cell(row_index, 3).value, worksheet.cell(row_index, 4).value,
				 worksheet.cell(row_index, 5).value, worksheet.cell(row_index, 6).value, worksheet.cell(row_index, 7).value, worksheet.cell(row_index, 8).value, worksheet.cell(row_index, 9).value,
				 worksheet.cell(row_index, 10).value, worksheet.cell(row_index, 11).value)
	
	cursor.execute(add_employee, data_employee)

def convert_date_format(current_date):
	#return datetime.datetime.strptime(current_date, '%Y%m%d').strftime('%Y-%m-%d')
	if current_date == "NULL":
		return "NULL"
	if type(current_date) is str:
		return datetime.datetime.strptime(str(current_date), '%d/%m/%Y').strftime('%Y-%m-%d')
	return current_date.strftime('%Y-%m-%d')
	
wrong_vehicle_id = [14886, 14892, 14894, 14963, 14988, 15021, 15145, 15206] 
	
worksheet = workbook.get_sheet_by_name(sheetnames[6]) # order					
add_salesorder = ("INSERT INTO salesorder "
				  "(orderID, createDate, pickupDate, pickupStoreID, returnDate, returnStoreID, customerID, vehicleID) "
				  "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)")
for row_index in range(2, 595):
	data_salesorder = (worksheet.cell(row_index, 1).value, convert_date_format(worksheet.cell(row_index, 2).value), convert_date_format(worksheet.cell(row_index, 3).value), worksheet.cell(row_index, 4).value,
				 convert_date_format(worksheet.cell(row_index, 5).value), worksheet.cell(row_index, 6).value, worksheet.cell(row_index, 7).value, worksheet.cell(row_index, 8).value)
	print (data_salesorder)
	if worksheet.cell(row_index, 4).value <= 40 and worksheet.cell(row_index, 6).value <= 40:
		if worksheet.cell(row_index, 8).value not in wrong_vehicle_id:
			cursor.execute(add_salesorder, data_salesorder)
cnx.commit()
cursor.close()
cnx.close()
